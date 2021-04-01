"""
作者:Jetery
日期:2021//04//01//
"""


"""题目要求
1. 通过代码，打开material文件夹下的文件practice2.xlsx，获取下半年公司名单工作表。
2. 打印第四列（D列）除表头部门外的所有数据。
3. 将原有的值全部修改为战略储备部。
4. 将结果保存为practice2_result.xlsx。
"""

from openpyxl import load_workbook

# 打开【practice2.xlsx】工作簿
staff_wb = load_workbook('./material/practice2.xlsx')
# 按表名取表
staff_ws = staff_wb['下半年公司名单']

# 循环获取第四列（D列）的所有单元格对象
for col_cell in staff_ws['D']:
    # 如果为表头，则跳过本次循环
    if col_cell.value == '部门':
        continue
    # 打印原有的值
        print(col_cell.value)
    # 将原有的值修改为'战略储备部'
        col_cell.value = '战略储备部'

# 将结果保存为【'practice2_result.xlsx'】
staff_wb.save('./material/practice2_result.xlsx')



"""
values_only为默认的False时，我们得到的row就是一个个由单元格对象组成的元组，可以通过索引或者for循环遍历的方式来获取单独的单元格对象。

还是使用上面的例子，不过这次我们不写参数values_only，让其保持为默认的False，这样iter_rows()就会返回指定的单元格对象了。

运行下面的代码，然后观察其结果与之前values_only=True时有何不同。
"""
#   # 打开【公司人员名单.xlsx】工作簿
#   staff_wb = load_workbook('./codes/material/公司人员名单.xlsx')
#   # 获取【'上半年公司名单'】工作表
#   fhy_ws = staff_wb['上半年公司名单']
#
#   # 返回第2行至第12行，第2列（B列）至第3列（C列）这个范围的所有单元格对象
#   for row in fhy_ws.iter_rows(min_row=2, max_row=12, min_col=2, max_col=3):
#       print(row)
"""

获取并打印了表格中第2行至第12行，第2列（B列）至第3列（C列）这个范围内的所有单元格对象。

虽然第十二行为空，但也可以获取到对应的单元格对象。

"""



"""

第二种方式和第一种大同小异，可以写为：

for cell in 工作表对象[行数]
for cell in 工作表对象['列名']

下面的代码分别取出了第三行和第三列的所有单元格对象，你可以运行后查看代码的结果。

"""

#   # 打开【公司人员名单.xlsx】工作簿
#   staff_wb = load_workbook('./codes/material/公司人员名单.xlsx')
#   # 获取活动工作表
#   staff_ws = staff_wb.active
#
#   # for循环遍历，取出第三行的所有单元格对象
#   for row_cell in staff_ws[3]:
#       print(row_cell)
#
#   # for循环遍历，取出第三列（C列）的所有单元格对象
#   for col_cell in staff_ws['C']:
#       print(col_cell)